from __future__ import annotations

import io
from pathlib import Path

CLAIMED_OFFICE_EXTENSIONS = frozenset({".docx", ".xlsx", ".pptx"})
CLAIMED_OFFICE_FORMATS = frozenset({"docx", "xlsx", "pptx"})
OFFICE_PREVIEW_KIND = "office"
OFFICE_RENDER_MODE = "code"
OFFICE_DEPENDENCY_HINT = (
    "Office preview is not available on this server. Install python-docx, "
    "openpyxl, and python-pptx to enable it: pip install python-docx openpyxl "
    "python-pptx"
)
OFFICE_PREVIEW_TRUNCATED_NOTICE = "[Preview truncated: Office content exceeds safe limits]"
MAX_OFFICE_PREVIEW_CHARS = 120_000
MAX_DOCX_PREVIEW_BLOCKS = 2_000
MAX_DOCX_TABLE_CELLS = 5_000
MAX_XLSX_PREVIEW_SHEETS = 20
MAX_XLSX_PREVIEW_ROWS_PER_SHEET = 500
MAX_XLSX_PREVIEW_CELLS_PER_SHEET = 5_000
MAX_PPTX_PREVIEW_SLIDES = 100
MAX_PPTX_PREVIEW_SHAPES_PER_SLIDE = 200

_WORD_NAMESPACE = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
_DEFAULT_DOCX_SECTION_SIGNATURE = None

_DOCX_BODY_CHILDREN = {f"{_WORD_NAMESPACE}p", f"{_WORD_NAMESPACE}sectPr"}
_DOCX_PARAGRAPH_CHILDREN = {f"{_WORD_NAMESPACE}pPr", f"{_WORD_NAMESPACE}r"}
_DOCX_SAFE_PARAGRAPH_PROPERTY_CHILDREN = {f"{_WORD_NAMESPACE}pStyle"}
_DOCX_RUN_CHILDREN = {f"{_WORD_NAMESPACE}t"}


def _office_dependency_import_error() -> ImportError:
    return ImportError(OFFICE_DEPENDENCY_HINT)


def _load_docx_document():
    try:
        from docx import Document as document_factory
    except ImportError as exc:  # pragma: no cover - depends on local install shape
        raise _office_dependency_import_error() from exc
    return document_factory


def _load_workbook_reader():
    try:
        from openpyxl import load_workbook as workbook_reader
    except ImportError as exc:  # pragma: no cover - depends on local install shape
        raise _office_dependency_import_error() from exc
    return workbook_reader


def _load_presentation_ctor():
    try:
        from pptx import Presentation as presentation_ctor
    except ImportError as exc:  # pragma: no cover - depends on local install shape
        raise _office_dependency_import_error() from exc
    return presentation_ctor


def is_claimed_office_path(path: str | Path) -> bool:
    return Path(str(path)).suffix.lower() in CLAIMED_OFFICE_EXTENSIONS


def _office_format_for_path(path: str | Path) -> str:
    return Path(str(path)).suffix.lower().lstrip(".")


def _normalise_preview_text(value, max_chars: int | None = None) -> str:
    if value is None:
        return ""
    text = str(value)
    if max_chars is not None and max_chars >= 0 and len(text) > max_chars:
        text = text[:max_chars]
    return text.replace("\r", "\n").replace("\n", " ").strip()


def _preview_line_count(content: str) -> int:
    if not content:
        return 1
    return content.count("\n") + 1


def _finalize_preview_text(content: str, truncated: bool = False) -> tuple[str, bool]:
    text = (content or "").strip()
    if len(text) > MAX_OFFICE_PREVIEW_CHARS:
        text = text[:MAX_OFFICE_PREVIEW_CHARS].rstrip()
        truncated = True
    if truncated:
        text = f"{text}\n\n{OFFICE_PREVIEW_TRUNCATED_NOTICE}" if text else OFFICE_PREVIEW_TRUNCATED_NOTICE
    return text, truncated


class _PreviewBuilder:
    def __init__(self, char_limit: int | None = None) -> None:
        self._char_limit = MAX_OFFICE_PREVIEW_CHARS if char_limit is None else max(char_limit, 0)
        self._parts: list[str] = []
        self._length = 0
        self._started = False
        self.truncated = False

    @property
    def remaining_chars(self) -> int:
        return max(self._char_limit - self._length, 0)

    @property
    def started(self) -> bool:
        return self._started

    @property
    def has_content(self) -> bool:
        return bool(self._parts)

    @property
    def text(self) -> str:
        return "".join(self._parts)

    def _append_piece(self, piece: str) -> bool:
        if self.truncated:
            return False
        if not piece:
            return True
        remaining = self.remaining_chars
        if remaining <= 0:
            self.truncated = True
            return False
        if len(piece) > remaining:
            piece = piece[:remaining].rstrip()
            self.truncated = True
        if piece:
            self._parts.append(piece)
            self._length += len(piece)
        return not self.truncated

    def start_line(self) -> bool:
        if self._started:
            return self._append_piece("\n")
        self._started = True
        return True

    def start_section(self) -> bool:
        if self._started:
            return self._append_piece("\n\n")
        self._started = True
        return True

    def append_text(self, text: str) -> bool:
        if not self._started:
            self._started = True
        return self._append_piece(text)

    def finish(self) -> tuple[str, bool]:
        return _finalize_preview_text(self.text, self.truncated)


def _append_normalized_preview_text(builder: _PreviewBuilder, value) -> bool:
    if builder.truncated:
        return False
    remaining = builder.remaining_chars
    if remaining <= 0:
        builder.truncated = True
        return False
    raw_text = "" if value is None else str(value)
    clipped = len(raw_text) > remaining
    if not builder.append_text(_normalise_preview_text(raw_text, remaining)):
        return False
    if clipped:
        builder.truncated = True
        return False
    return True


def _iter_docx_text_nodes(element):
    for node in element.iter():
        if node.tag == f"{_WORD_NAMESPACE}t" and node.text:
            yield node.text


def _append_docx_element_text(builder: _PreviewBuilder, element) -> bool:
    for text in _iter_docx_text_nodes(element):
        if not _append_normalized_preview_text(builder, text):
            return False
    return True


def _append_docx_cell_text(builder: _PreviewBuilder, cell_element) -> bool:
    first_paragraph = True
    for child in cell_element:
        if child.tag != f"{_WORD_NAMESPACE}p":
            continue
        if not first_paragraph and not builder.append_text("\n"):
            return False
        if not _append_docx_element_text(builder, child):
            return False
        first_paragraph = False
    return True


def _docx_preview_text(document) -> tuple[str, bool]:
    builder = _PreviewBuilder()
    body_blocks_seen = 0
    table_cells_seen = 0
    table_index = 0
    for child in document._element.body:
        if child.tag == f"{_WORD_NAMESPACE}sectPr":
            continue
        body_blocks_seen += 1
        if body_blocks_seen > MAX_DOCX_PREVIEW_BLOCKS:
            builder.truncated = True
            break
        if child.tag == f"{_WORD_NAMESPACE}p":
            if not builder.start_line() or not _append_docx_element_text(builder, child):
                break
            continue
        if child.tag != f"{_WORD_NAMESPACE}tbl":
            continue
        table_index += 1
        if not builder.start_line() or not builder.append_text(f"Table {table_index}"):
            break
        for row in child:
            if row.tag != f"{_WORD_NAMESPACE}tr":
                continue
            if not builder.start_line():
                break
            first_cell = True
            for cell in row:
                if cell.tag != f"{_WORD_NAMESPACE}tc":
                    continue
                table_cells_seen += 1
                if table_cells_seen > MAX_DOCX_TABLE_CELLS:
                    builder.truncated = True
                    break
                if not first_cell and not builder.append_text("\t"):
                    break
                if not _append_docx_cell_text(builder, cell):
                    break
                first_cell = False
            if builder.truncated:
                break
        if builder.truncated:
            break
    return builder.finish()


def _docx_paragraph_properties_are_safe(properties) -> bool:
    for child in properties:
        if child.tag not in _DOCX_SAFE_PARAGRAPH_PROPERTY_CHILDREN:
            return False
        if child.tag == f"{_WORD_NAMESPACE}pStyle" and child.get(f"{_WORD_NAMESPACE}val") != "Normal":
            return False
    return True


def _docx_xml_signature(element) -> tuple:
    attributes = tuple(
        sorted(
            (key, value)
            for key, value in element.attrib.items()
            if not key.rsplit("}", 1)[-1].startswith("rsid")
        )
    )
    children = tuple(_docx_xml_signature(child) for child in element)
    text = (element.text or "").strip()
    return element.tag, attributes, text, children


def _default_docx_section_signature() -> tuple:
    global _DEFAULT_DOCX_SECTION_SIGNATURE
    if _DEFAULT_DOCX_SECTION_SIGNATURE is None:
        document = _load_docx_document()()
        _DEFAULT_DOCX_SECTION_SIGNATURE = tuple(
            _docx_xml_signature(child) for child in document._element.body.sectPr
        )
    return _DEFAULT_DOCX_SECTION_SIGNATURE


def _docx_section_properties_are_safe(section_properties) -> bool:
    return tuple(_docx_xml_signature(child) for child in section_properties) == _default_docx_section_signature()


def _docx_editability(document) -> tuple[bool, str | None]:
    body = document._element.body
    for child in body:
        if child.tag not in _DOCX_BODY_CHILDREN:
            return False, "docx contains unsupported structures"
        if child.tag == f"{_WORD_NAMESPACE}sectPr" and not _docx_section_properties_are_safe(child):
            return False, "docx contains unsupported section content"
    for paragraph in document.paragraphs:
        for child in paragraph._p:
            if child.tag not in _DOCX_PARAGRAPH_CHILDREN:
                return False, "docx contains unsupported paragraph structures"
            if child.tag == f"{_WORD_NAMESPACE}pPr" and not _docx_paragraph_properties_are_safe(child):
                return False, "docx contains unsupported paragraph structures"
        for run in paragraph.runs:
            for child in run._r:
                if child.tag not in _DOCX_RUN_CHILDREN:
                    return False, "docx contains unsupported inline content"
    return True, None


def _preview_docx(raw: bytes) -> tuple[str, bool, str | None, bool]:
    try:
        document = _load_docx_document()(io.BytesIO(raw))
    except ImportError:
        raise
    except Exception as exc:  # pragma: no cover - library-specific failure mode
        raise ValueError("Unable to read DOCX preview") from exc
    content, truncated = _docx_preview_text(document)
    if truncated:
        return content, False, "docx preview exceeds safe limits", True
    editable, reason = _docx_editability(document)
    return content, editable, reason, truncated


def _preview_xlsx(raw: bytes) -> tuple[str, bool]:
    try:
        workbook = _load_workbook_reader()(io.BytesIO(raw), data_only=True, read_only=True)
    except ImportError:
        raise
    except Exception as exc:  # pragma: no cover - library-specific failure mode
        raise ValueError("Unable to read XLSX preview") from exc
    builder = _PreviewBuilder()
    try:
        for sheet_index, sheet in enumerate(workbook.worksheets, start=1):
            if sheet_index > MAX_XLSX_PREVIEW_SHEETS:
                builder.truncated = True
                break
            if not builder.start_section() or not builder.append_text(f"Sheet: {sheet.title}"):
                break
            rows_seen = 0
            cells_seen = 0
            max_row = min(getattr(sheet, "max_row", MAX_XLSX_PREVIEW_ROWS_PER_SHEET), MAX_XLSX_PREVIEW_ROWS_PER_SHEET)
            max_col = min(getattr(sheet, "max_column", MAX_XLSX_PREVIEW_CELLS_PER_SHEET), MAX_XLSX_PREVIEW_CELLS_PER_SHEET)
            for row in sheet.iter_rows(values_only=True, max_row=max_row, max_col=max_col):
                rows_seen += 1
                row_budget = builder.remaining_chars - (1 if builder.started else 0)
                row_builder = _PreviewBuilder(row_budget)
                for value in row:
                    cells_seen += 1
                    if cells_seen > MAX_XLSX_PREVIEW_CELLS_PER_SHEET:
                        builder.truncated = True
                        break
                    if row_builder.has_content and not row_builder.append_text("\t"):
                        break
                    if not _append_normalized_preview_text(row_builder, value):
                        if not row_builder.truncated:
                            builder.truncated = True
                        break
                if builder.truncated:
                    break
                if row_builder.has_content:
                    if not builder.start_line() or not builder.append_text(row_builder.text):
                        break
                    if row_builder.truncated:
                        builder.truncated = True
                        break
            if builder.truncated:
                break
            if getattr(sheet, "max_row", rows_seen) > MAX_XLSX_PREVIEW_ROWS_PER_SHEET:
                builder.truncated = True
                break
            if getattr(sheet, "max_column", 0) > MAX_XLSX_PREVIEW_CELLS_PER_SHEET:
                builder.truncated = True
                break
    finally:
        close = getattr(workbook, "close", None)
        if callable(close):
            close()
    if not builder.has_content:
        return _finalize_preview_text("Empty workbook", builder.truncated)
    return builder.finish()


def _preview_pptx(raw: bytes) -> tuple[str, bool]:
    try:
        presentation = _load_presentation_ctor()(io.BytesIO(raw))
    except ImportError:
        raise
    except Exception as exc:  # pragma: no cover - library-specific failure mode
        raise ValueError("Unable to read PPTX preview") from exc
    builder = _PreviewBuilder()
    for slide_index, slide in enumerate(presentation.slides, start=1):
        if slide_index > MAX_PPTX_PREVIEW_SLIDES:
            builder.truncated = True
            break
        if not builder.start_section() or not builder.append_text(f"Slide {slide_index}"):
            break
        shapes_seen = 0
        has_text = False
        for shape in slide.shapes:
            shapes_seen += 1
            if shapes_seen > MAX_PPTX_PREVIEW_SHAPES_PER_SLIDE:
                builder.truncated = True
                break
            text = getattr(shape, "text", "")
            if not _normalise_preview_text(text):
                continue
            if not builder.start_line() or not _append_normalized_preview_text(builder, text):
                break
            has_text = True
        if builder.truncated:
            break
        if not has_text and (not builder.start_line() or not builder.append_text("(empty slide)")):
            break
    if not builder.has_content:
        return _finalize_preview_text("Empty presentation", builder.truncated)
    return builder.finish()


def preview_office_document(path: str | Path, raw: bytes) -> dict:
    office_format = _office_format_for_path(path)
    if office_format not in CLAIMED_OFFICE_FORMATS:
        raise ValueError(f"Unsupported Office format: {path}")

    truncated = False
    if office_format == "docx":
        content, editable, reason, truncated = _preview_docx(raw)
    elif office_format == "xlsx":
        content, truncated = _preview_xlsx(raw)
        editable, reason = False, "xlsx preview is read-only in this slice"
    elif office_format == "pptx":
        content, truncated = _preview_pptx(raw)
        editable, reason = False, "pptx preview is read-only in this slice"
    else:  # pragma: no cover - exhaustive guard
        raise ValueError(f"Unsupported Office format: {path}")

    payload = {
        "path": str(path),
        "content": content,
        "size": len(raw),
        "lines": _preview_line_count(content),
        "preview_kind": OFFICE_PREVIEW_KIND,
        "office_format": office_format,
        "render_mode": OFFICE_RENDER_MODE,
        "editable": editable,
    }
    if reason:
        payload["edit_blocked_reason"] = reason
    if truncated:
        payload["truncated"] = True
    return payload


def _docx_bytes_from_text(content: str) -> bytes:
    document = _load_docx_document()()
    body = document._element.body
    for child in list(body):
        if child.tag != f"{_WORD_NAMESPACE}sectPr":
            body.remove(child)
    text = str(content or "").replace("\r\n", "\n").replace("\r", "\n")
    for line in text.split("\n"):
        document.add_paragraph(line)
    buffer = io.BytesIO()
    document.save(buffer)
    return buffer.getvalue()


def save_office_document(path: str | Path, current_bytes: bytes, content: str) -> tuple[dict, bytes]:
    office_format = _office_format_for_path(path)
    if office_format != "docx":
        raise ValueError(f"{office_format or 'office file'} is preview-only in this slice")

    current_preview = preview_office_document(path, current_bytes)
    if not current_preview.get("editable"):
        raise ValueError(current_preview.get("edit_blocked_reason") or "DOCX document is not editable")

    saved_bytes = _docx_bytes_from_text(content)
    saved_preview = preview_office_document(path, saved_bytes)
    if not saved_preview.get("editable"):
        raise ValueError(saved_preview.get("edit_blocked_reason") or "Saved DOCX is not editable")
    return saved_preview, saved_bytes
